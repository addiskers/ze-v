"""
Contacts import helpers for the EO Admin platform.

- `normalize_phone` — India-centric E.164 normaliser that rejects the classic
  Excel scientific-notation corruption (e.g. `9.17619E+11`).
- `parse_upload` — read an .xlsx (openpyxl) or .csv into (name, e164, status)
  rows, de-duplicated by phone within the batch.
- `build_template` — a real .xlsx sample with a Text-formatted Phone column so
  users don't re-introduce scientific notation.
"""

import csv
import io
import re

from openpyxl import Workbook, load_workbook

_MAX_ROWS = 100_000
_NAME_HEADERS = {"name", "full name", "contact name", "contact", "member", "guest"}
_PHONE_HINTS = ("phone", "mobile", "number", "contact no", "whatsapp", "cell", "msisdn")


def normalize_phone(raw):
    """Return (e164_or_None, is_valid). None => unparseable / rejected."""
    if raw is None:
        return None, False
    s = str(raw).strip()
    if not s:
        return None, False
    # reject float / scientific-notation corruption ("9.17619E+11", "917619000000.0")
    if re.search(r"[eE][+\-]?\d", s) or re.fullmatch(r"\d+\.\d+", s):
        return None, False
    plus = s.startswith("+")
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None, False
    if plus:
        e164 = "+" + digits
    elif len(digits) == 10:
        e164 = "+91" + digits                       # bare Indian mobile
    elif len(digits) == 12 and digits.startswith("91"):
        e164 = "+" + digits
    elif len(digits) == 11 and digits.startswith("0"):
        e164 = "+91" + digits[1:]                    # leading-0 domestic form
    else:
        e164 = "+" + digits
    ndig = len(e164) - 1
    return e164, (10 <= ndig <= 15)


def _pick_columns(header):
    name_idx = phone_idx = None
    for i, cell in enumerate(header):
        c = str(cell or "").strip().lower()
        if name_idx is None and c in _NAME_HEADERS:
            name_idx = i
        if phone_idx is None and any(k in c for k in _PHONE_HINTS):
            phone_idx = i
    return name_idx, phone_idx


def _rows_from_matrix(matrix):
    """matrix: list of row-tuples. Returns list of (name_raw, phone_raw)."""
    matrix = [r for r in matrix if r is not None and any(c not in (None, "") for c in r)]
    if not matrix:
        return []
    name_idx, phone_idx = _pick_columns(matrix[0])
    if phone_idx is not None:
        body = matrix[1:]                            # first row was a header
    else:
        ncol = max(len(r) for r in matrix)
        name_idx, phone_idx = (0, 1) if ncol >= 2 else (None, 0)
        body = matrix
    out = []
    for r in body:
        ph = r[phone_idx] if phone_idx is not None and phone_idx < len(r) else None
        nm = r[name_idx] if name_idx is not None and name_idx < len(r) else None
        if ph in (None, "") and nm in (None, ""):
            continue
        out.append((nm, ph))
    return out


def _parse_xlsx(data):
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    matrix = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= _MAX_ROWS:
            break
        matrix.append(row)
    wb.close()
    return _rows_from_matrix(matrix)


def _parse_csv(data):
    text = data.decode("utf-8-sig", errors="replace")
    matrix = [tuple(r) for r in csv.reader(io.StringIO(text))]
    return _rows_from_matrix(matrix[:_MAX_ROWS])


def parse_upload(filename, data):
    """Return (rows, rejected, total). rows: list of (name, e164, status)."""
    name = (filename or "").lower()
    raw_rows = _parse_csv(data) if name.endswith(".csv") else _parse_xlsx(data)
    seen = {}                                        # e164 -> (name, status)
    rejected = 0
    for nm, ph in raw_rows:
        e164, valid = normalize_phone(ph)
        if not e164:
            rejected += 1
            continue
        nm = (str(nm).strip() if nm not in (None, "") else "")
        status = "valid" if valid else "invalid"
        prev = seen.get(e164)
        # keep a name if we have one; prefer valid status
        keep_name = nm or (prev[0] if prev else "")
        keep_status = "valid" if (status == "valid" or (prev and prev[1] == "valid")) else "invalid"
        seen[e164] = (keep_name, keep_status)
    rows = [(nm, ph, st) for ph, (nm, st) in seen.items()]
    return rows, rejected, len(raw_rows)


def build_template():
    """A minimal .xlsx sample: headers Name/Phone, phone column Text-formatted."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws["A1"] = "Name"
    ws["B1"] = "Phone"
    samples = [("Rahul Shah", "9876543210"), ("Priya Patel", "+919812345678")]
    for r, (nm, ph) in enumerate(samples, start=2):
        ws.cell(row=r, column=1, value=nm)
        c = ws.cell(row=r, column=2, value=ph)
        c.number_format = "@"                        # Text — preserves leading digits
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
